import io
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
import openpyxl
from docx import Document

def export_villes_to_pdf(villes):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    data = [["Nom", "Région"]]
    for ville in villes:
        data.append([ville.nom, ville.region.nom])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    doc.build([table])
    buffer.seek(0)
    return buffer

def export_villes_to_excel(villes):
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws.append(["Nom", "Région"])
    for ville in villes:
        ws.append([ville.nom, ville.region.nom])
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_villes_to_word(villes):
    buffer = io.BytesIO()
    doc = Document()
    
    table = doc.add_table(rows=1, cols=2)
    table.style = 'Table Grid'
    hdr_cells = table.rows[0].cells
    hdr_cells[0].text = 'Nom'
    hdr_cells[1].text = 'Région'
    
    for ville in villes:
        row_cells = table.add_row().cells
        row_cells[0].text = ville.nom
        row_cells[1].text = ville.region.nom
    
    doc.save(buffer)
    buffer.seek(0)
    return buffer

def export_regions_to_excel(regions):
    """Exporte une liste de régions vers un fichier Excel."""
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Régions"
    
    # En-têtes
    headers = ["ID", "Nom de la région", "Nombre de villes"]
    ws.append(headers)
    
    # Données
    for region in regions:
        ws.append([
            region.id,
            region.nom,
            region.ville_set.count()  # Compte le nombre de villes associées
        ])
    
    # Mise en forme
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Style de l'en-tête
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="DDDDDD")
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_regions_to_pdf(regions):
    """Exporte une liste de régions vers un fichier PDF."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    # Préparation des données
    data = [["ID", "Nom de la région", "Villes"]]
    
    for region in regions:
        data.append([
            str(region.id),
            region.nom,
            str(region.ville_set.count())
        ])
    
    # Création du tableau
    table = Table(data)
    
    # Style du tableau
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8f9fa')),  # En-tête gris clair
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#212529')),  # Texte noir
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Fond blanc pour les données
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#495057')),  # Texte gris foncé
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),  # Bordures gris clair
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),  # Bordure extérieure
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    
    # Appliquer le style
    table.setStyle(style)
    
    # Ajuster la largeur des colonnes
    table._argW[0] = 2*cm  # ID
    table._argW[1] = 8*cm  # Nom
    table._argW[2] = 4*cm  # Nombre de villes
    
    # Construire le document
    elements = [table]
    doc.build(elements)
    
    buffer.seek(0)
    return buffer